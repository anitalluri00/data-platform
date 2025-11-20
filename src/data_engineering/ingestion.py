import pandas as pd
import requests
from bs4 import BeautifulSoup
import PyPDF2
from docx import Document
import openpyxl
import csv
import os
from typing import Union, Dict, Any
import logging
from database.operations import DatabaseOperations

logger = logging.getLogger(__name__)

class DataIngestion:
    def __init__(self):
        self.db_ops = DatabaseOperations()
    
    def ingest_file(self, file_path: str, source_name: str) -> int:
        """Ingest various file types and return source ID"""
        try:
            file_type = file_path.split('.')[-1].lower()
            file_size = os.path.getsize(file_path)
            
            # Store source info
            source_id = self.db_ops.insert_data_source(
                source_name=source_name,
                source_type='file',
                file_path=file_path,
                file_type=file_type,
                file_size=file_size
            )
            
            # Process based on file type
            if file_type in ['txt', 'rtf']:
                self._process_text_file(file_path, source_id)
            elif file_type in ['pdf']:
                self._process_pdf(file_path, source_id)
            elif file_type in ['docx', 'doc']:
                self._process_docx(file_path, source_id)
            elif file_type in ['xlsx', 'xls']:
                self._process_excel(file_path, source_id)
            elif file_type in ['csv']:
                self._process_csv(file_path, source_id)
            elif file_type in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                self._process_image(file_path, source_id)
            else:
                self._process_binary_file(file_path, source_id)
            
            logger.info(f"Successfully ingested file: {file_path}")
            return source_id
            
        except Exception as e:
            logger.error(f"Error ingesting file {file_path}: {str(e)}")
            raise
    
    def _process_text_file(self, file_path: str, source_id: int):
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        self.db_ops.insert_raw_data(source_id, 'text', content.encode(), content)
    
    def _process_pdf(self, file_path: str, source_id: int):
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        self.db_ops.insert_raw_data(source_id, 'pdf', file_content, text)
    
    def _process_docx(self, file_path: str, source_id: int):
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        self.db_ops.insert_raw_data(source_id, 'docx', file_content, text)
    
    def _process_excel(self, file_path: str, source_id: int):
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                text += ",".join([str(cell) if cell else "" for cell in row]) + "\n"
        
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        self.db_ops.insert_raw_data(source_id, 'excel', file_content, text)
    
    def _process_csv(self, file_path: str, source_id: int):
        df = pd.read_csv(file_path)
        text = df.to_string()
        
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        self.db_ops.insert_raw_data(source_id, 'csv', file_content, text)
    
    def _process_image(self, file_path: str, source_id: int):
        # For images, we'll store metadata and use OCR in processing stage
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        metadata = {
            'file_type': 'image',
            'file_size': os.path.getsize(file_path),
            'dimensions': 'To be extracted in processing'
        }
        
        self.db_ops.insert_raw_data(source_id, 'image', file_content, "", metadata)
    
    def _process_binary_file(self, file_path: str, source_id: int):
        with open(file_path, 'rb') as file:
            file_content = file.read()
        
        self.db_ops.insert_raw_data(source_id, 'binary', file_content, "")
    
    def ingest_web_content(self, url: str, source_name: str) -> int:
        """Ingest content from web URL"""
        try:
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            text = soup.get_text()
            
            source_id = self.db_ops.insert_data_source(
                source_name=source_name,
                source_type='web',
                source_url=url
            )
            
            self.db_ops.insert_raw_data(source_id, 'web', response.content, text)
            
            return source_id
            
        except Exception as e:
            logger.error(f"Error ingesting web content {url}: {str(e)}")
            raise